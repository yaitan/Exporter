from typing import Dict

import requests
import openpyxl as xl
import datetime

API = "https://graph.instagram.com/v23.0"
media_fields = ['id',
                'caption',
                'media_product_type',
                'media_type',
                'permalink',
                'timestamp',
                'like_count',
                'comments_count',
                'is_shared_to_feed'
                ]

comment_fields = ['text',
                  'id',
                  'like_count',
                  'replies',
                  'timestamp',
                  'parent_id',
                  'hidden',
                  ]


def add_collaborators(workbook: xl.Workbook):
    sheet = workbook.active
    sheet.cell(row=1, column=sheet.max_column + 1, value='collaborators')
    for row in sheet.iter_rows(min_row=2):
        media_id = row[0].value
        collabs_json = get_collabs(media_id)
        collabs = ', '.join([data['username'] for data in collabs_json['data']])
        row[-1].value = collabs


def get_collabs(media_id):
    url = f"{API}/{media_id}/collaborators&{ACCESS}"

    response = requests.get(url)
    return response.json()


def reformat_media_product_type(json):
    product = json['media_product_type']
    if product == "REELS":
        feed = json['is_shared_to_feed']
        json['media_product_type'] = 'Reel in feed' if feed else 'Reel'
    else:
        json['media_product_type'] = 'Post'
    return json


def create_xl(name):
    wb = xl.Workbook()
    sheet = wb.active
    sheet.title = "Posts"
    sheet.append(media_fields[:-1] + ['collaborators'])
    for column in sheet.columns:
        sheet.column_dimensions[column[0].column_letter].width = 20
    wb.save(name)
    return wb


def edit_margins(workbook):
    sheet = workbook.active
    for column in sheet.columns:
        sheet.column_dimensions[column[0].column_letter].width = 20


def open_xl(filename):
    workbook = xl.load_workbook(filename)
    return workbook


def add_all_media(media_list):
    i = 1
    sheet = workbook.active
    count = 0
    for media in media_list:
        data = get_media_info(media['id'])
        count += data['comments_count']
        data['timestamp'] = reformat_time(data['timestamp'])
        # collabs_json = get_collabs(media['id'])
        # collabs = ', '.join([data['username'] for data in collabs_json['data']])
        # row.append(collabs)
        sheet.append(list(data.values())[:-1])

        print('\r', end='')
        print(f'{i}/{len(media_list)}', end='')
        i += 1
    print("\nAdded media")
    workbook.save(filename)
    return count


def reformat_time(time):
    dt = datetime.datetime.strptime(time, '%Y-%m-%dT%H:%M:%S%z')
    excel_format = dt.strftime('%Y-%m-%d %H:%M:%S')
    return excel_format


def create_comments_sheet():
    sheet = workbook.create_sheet("Comments")
    sheet.append(['post'] + comment_fields)
    for column in sheet.columns:
        sheet.column_dimensions[column[0].column_letter].width = 20
    workbook.save(filename)
    return sheet


def standarize_comment(comment: Dict, media_id):
    standard_fields = {'id': None, 'post_id': media_id, 'text': None, 'like_count': 0, 'replies': {'data': []},
                       'timestamp': None, 'parent_id': None, 'hidden': 'UNKNOWN'}
    stand_comment = {k: comment.get(k, v) for k, v in standard_fields.items()}
    stand_comment['replies'] = len(stand_comment['replies']['data'])
    stand_comment['timestamp'] = reformat_time(stand_comment['timestamp'])
    return stand_comment


def add_comments(media_id, sheet):
    comments = get_comments(media_id)
    for comment in comments:
        comment = standarize_comment(comment, media_id)
        sheet.append(list(comment.values()))
    workbook.save(filename)
    return len(comments)


def get_username():
    url = f"{API}/me?fields=user_id,username&{ACCESS}"
    username = requests.get(url)
    return username.json()["user_id"]


def get_post_count(ig_user_id):
    url = f"{API}/{ig_user_id}?fields=media_count&{ACCESS}"
    return requests.get(url).json()["media_count"]


# Step 1: Fetch media
def get_media_ids(ig_user_id):
    data = []
    url = f"{API}/{ig_user_id}/media?{ACCESS}&limit=100"
    response = requests.get(url)
    json = response.json()
    data += json['data']
    print(f'fetched {len(data)} posts')

    while 'next' in json['paging']:
        response = requests.get(json['paging']['next'])
        json = response.json()
        data += json['data']
        print(f'fetched {len(data)} posts')
    return data


# Step 2: Fetch comments for a specific media ID
def get_comments(media_id):
    data = []
    fields = ','.join(comment_fields)
    url = f"{API}/{media_id}/comments?fields={fields}&{ACCESS}&limit=100"
    response = requests.get(url)
    json = response.json()
    data += json.get('data', 0)
    while 'next' in json.get('paging', []):
        response = requests.get(json['paging']['next'])
        json = response.json()
        data += json['data']
    return data


def get_media_info(media_id):
    fields = ",".join(media_fields)
    url = f"{API}/{media_id}/?fields={fields}&{ACCESS}"
    response = requests.get(url)
    json = reformat_media_product_type(response.json())
    return json


# Example Usage
# media = get_media(INSTAGRAM_USER_ID, ACCESS_TOKEN)
# print("Media:", media)

# if "data" in media and media["data"]:
# first_media_id = media["data"][0]["id"]  # Fetch the first media ID
# comments = get_comments(first_media_id, ACCESS_TOKEN)
# print("Comments:", comments)

# user_id = get_username()
# total_posts = get_post_count(user_id)
#

# test_media_id = '18050673620107822'
# user_id = get_username()
# total_posts = get_post_count(user_id)
# medias = get_media_ids(user_id)
# print(get_collabs(test_media_id))
# print(medias[0:10])
# comments = get_comments(medias[0]['id'])
# print(comments)
# sheets = create_xl("Test")
# add_all_media(medias, sheets)
# sheets.save("Test.xlsx")


if __name__ == '__main__':
    with open('key', 'r') as file:
        access_token = file.read()

    ACCESS = f"access_token={access_token}"

    user_id = get_username()
    medias = get_media_ids(user_id)
    date = datetime.datetime.now().strftime("%Y-%m-%d")
    filename = f"Snapshot_{date}.xlsx"
    workbook = create_xl(filename)
    comment_count = add_all_media(medias)
    comments_read = 0

    create_comments_sheet()
    for m in medias:
        comments_read += add_comments(m['id'], workbook['Comments'])
        print('\r', end='')
        print(f'Fetched {comments_read} comments out of {comment_count}', end='')
    print('\rFinished')
    workbook.save(filename)
